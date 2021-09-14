#input関数を使う

from openpyxl import load_workbook

filename = input('読み込むブック名(.xlsxまで全て)：')
cellno = input('読み込むセル名（例 A1）：')

wb = load_workbook(filename, read_only=True)
ws = wb.active

print(ws[cellno].value)