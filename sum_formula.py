#数式をセルに仕込む

from openpyxl import load_workbook

wb = load_workbook("作業時間.xlsx")
ws = wb.active

ws["D2"] = "=SUM(D4:D13)"

wb.save("作業時間_集計後.slsx")